#!/usr/bin/env python3
"""
Markdown to Excel Datasheet Converter

Converts markdown template files to professional Excel datasheets.

Usage:
    python scripts/md_to_excel.py templates/101-GR.md
    python scripts/md_to_excel.py --all
    python scripts/md_to_excel.py --all --output-dir assets/
"""

import argparse
import json
import re
import sys
from pathlib import Path
from typing import Any

import yaml

try:
    from openpyxl import Workbook
    from openpyxl.worksheet.datavalidation import DataValidation
except ImportError:
    print("Error: openpyxl is required. Install with: pip install openpyxl")
    sys.exit(1)

# Import centralized styles
from excel_styles import (
    FONT_TITLE, FONT_SECTION, FONT_LABEL, FONT_VALUE, FONT_VALUE_BOLD,
    FONT_SMALL, FONT_TINY,
    BORDER_ALL, NO_FILL, DOT_BORDER, THIN_BORDER,
    ALIGN_LEFT, ALIGN_CENTER, ALIGN_RIGHT,
    COLUMN_WIDTHS, ROW_HEIGHT,
    apply_print_layout, apply_column_widths, apply_row_heights,
    style_cell, apply_borders_to_range, apply_professional_borders,
    apply_section_borders
)
from openpyxl.styles import Border


# =============================================================================
# MARKDOWN PARSING
# =============================================================================

def parse_frontmatter(content: str) -> tuple[dict, str]:
    """Extract YAML frontmatter and remaining content."""
    if not content.startswith('---'):
        return {}, content

    # Find the closing ---
    end_match = re.search(r'\n---\s*\n', content[3:])
    if not end_match:
        return {}, content

    yaml_content = content[3:3 + end_match.start()]
    remaining = content[3 + end_match.end():]

    try:
        frontmatter = yaml.safe_load(yaml_content)
        return frontmatter or {}, remaining
    except yaml.YAMLError as e:
        print(f"Warning: Failed to parse YAML frontmatter: {e}")
        return {}, content


def parse_markdown_table(table_text: str) -> list[dict]:
    """Parse a markdown table into a list of row dictionaries."""
    lines = [line.strip() for line in table_text.strip().split('\n') if line.strip()]

    if len(lines) < 2:
        return []

    # Parse header row
    header_line = lines[0]
    headers = [h.strip() for h in header_line.split('|') if h.strip()]

    # Skip separator line (contains ---)
    rows = []
    for line in lines[2:]:
        if '---' in line:
            continue
        cells = [c.strip() for c in line.split('|') if c.strip() or line.count('|') > len(headers)]

        # Handle empty cells from split
        raw_cells = line.split('|')[1:-1] if line.startswith('|') else line.split('|')
        cells = [c.strip() for c in raw_cells]

        if len(cells) >= len(headers):
            row = {headers[i]: cells[i] if i < len(cells) else '' for i in range(len(headers))}
            rows.append(row)

    return rows


def extract_sections(content: str) -> dict[str, Any]:
    """Extract sections from markdown content."""
    sections = {}

    # Split by ## headers
    section_pattern = r'^##\s+(.+?)\s*$'
    parts = re.split(section_pattern, content, flags=re.MULTILINE)

    # parts[0] is content before first ##
    # parts[1], parts[3], parts[5]... are section names
    # parts[2], parts[4], parts[6]... are section contents

    for i in range(1, len(parts), 2):
        section_name = parts[i].strip()
        section_content = parts[i + 1] if i + 1 < len(parts) else ''
        sections[section_name] = section_content.strip()

    return sections


def parse_template(filepath: Path) -> dict:
    """Parse a markdown template file into structured data."""
    content = filepath.read_text(encoding='utf-8')

    frontmatter, body = parse_frontmatter(content)
    sections = extract_sections(body)

    template = {
        'metadata': frontmatter,
        'sections': {}
    }

    # Parse each section's table
    for section_name, section_content in sections.items():
        # Find table in section content
        table_match = re.search(r'(\|.+\|[\s\S]*?)(?=\n\n|\n##|\Z)', section_content)
        if table_match:
            table_text = table_match.group(1)
            template['sections'][section_name] = parse_markdown_table(table_text)
        else:
            template['sections'][section_name] = []

    return template


# =============================================================================
# EXCEL GENERATION
# =============================================================================

def parse_options(options_str: str) -> list[str]:
    """Parse dropdown options from JSON array or comma-separated string."""
    if not options_str:
        return []

    options_str = options_str.strip()

    # Try JSON array first
    if options_str.startswith('['):
        try:
            return json.loads(options_str)
        except json.JSONDecodeError:
            pass

    # Fall back to comma-separated
    return [opt.strip() for opt in options_str.split(',') if opt.strip()]


def build_workbook(template: dict) -> Workbook:
    """Build an Excel workbook from parsed template data."""
    wb = Workbook()
    ws = wb.active

    metadata = template['metadata']
    sections = template['sections']

    # Set sheet name
    title = metadata.get('title', 'DATASHEET')
    ws.title = title[:31]  # Excel sheet name limit

    # Apply layout
    apply_column_widths(ws)
    apply_row_heights(ws, 1, 65)

    # Track current row
    row = 1

    # Track data validations to add at end
    validations = []

    # === TITLE BLOCK (Rows 1-2) ===
    ws.merge_cells(start_row=1, start_column=1, end_row=2, end_column=10)
    title_cell = ws.cell(row=1, column=1, value="TECHNICAL SPECIFICATION DATA SHEET")
    style_cell(title_cell, 'title')
    apply_borders_to_range(ws, 1, 2, 1, 10)
    row = 3

    # === PROFESSIONAL HEADER BLOCK (Rows 3-7) ===
    service_info = sections.get('Service Information', [])
    doc_control = sections.get('Document Control', [])

    # Extract document control values
    project_name = ''
    project_no = ''
    equipment_tag = ''
    doc_no = ''
    revision = ''
    doc_date = ''

    for item in doc_control:
        field = item.get('Field', '')
        value = item.get('Value', '')
        if 'Project' in field and 'No' not in field:
            project_name = value
        elif 'Tag' in field:
            equipment_tag = value
        elif 'Document' in field or 'Doc' in field:
            doc_no = value
        elif 'Revision' in field or 'Rev' in field:
            revision = value
        elif 'Date' in field:
            doc_date = value

    # Extract service info values
    service_type = metadata.get('service', '')
    p_and_id = ''

    for item in service_info:
        field = item.get('Field', '')
        value = item.get('Value', '')
        if 'Service' in field:
            service_type = value or service_type
        elif 'P&ID' in field:
            p_and_id = value

    # Helper for header row borders
    def header_left_border():
        return Border(left=THIN_BORDER, top=THIN_BORDER, bottom=THIN_BORDER)

    def header_right_border():
        return Border(right=THIN_BORDER, top=THIN_BORDER, bottom=THIN_BORDER)

    def header_middle_border():
        return Border(top=THIN_BORDER, bottom=THIN_BORDER)

    # --- Row 3: Project: [value] | DES BY: | SPEC NO: ---
    # Label A-B merged
    ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=2)
    ws.cell(row=3, column=1, value='Project:').font = FONT_LABEL
    ws.cell(row=3, column=1).border = header_left_border()
    ws.cell(row=3, column=2).border = header_right_border()

    # Value C-F merged
    ws.merge_cells(start_row=3, start_column=3, end_row=3, end_column=6)
    ws.cell(row=3, column=3, value=project_name).font = FONT_VALUE
    ws.cell(row=3, column=3).border = header_left_border()
    for col in range(4, 6):
        ws.cell(row=3, column=col).border = header_middle_border()
    ws.cell(row=3, column=6).border = header_right_border()

    # DES BY: G, value H
    ws.cell(row=3, column=7, value='DES BY:').font = FONT_TINY
    ws.cell(row=3, column=7).border = BORDER_ALL
    ws.cell(row=3, column=8).font = FONT_VALUE
    ws.cell(row=3, column=8).border = BORDER_ALL

    # SPEC NO: I, value J
    ws.cell(row=3, column=9, value='SPEC NO:').font = FONT_TINY
    ws.cell(row=3, column=9).border = BORDER_ALL
    ws.cell(row=3, column=10, value=doc_no).font = FONT_VALUE
    ws.cell(row=3, column=10).border = BORDER_ALL

    # --- Row 4: Project No. [value] | CHK BY: | REV: ---
    ws.merge_cells(start_row=4, start_column=1, end_row=4, end_column=2)
    ws.cell(row=4, column=1, value='Project No.').font = FONT_LABEL
    ws.cell(row=4, column=1).border = BORDER_ALL
    ws.cell(row=4, column=2).border = BORDER_ALL

    ws.merge_cells(start_row=4, start_column=3, end_row=4, end_column=6)
    ws.cell(row=4, column=3, value=project_no).font = FONT_VALUE
    ws.cell(row=4, column=3).border = header_left_border()
    for col in range(4, 6):
        ws.cell(row=4, column=col).border = header_middle_border()
    ws.cell(row=4, column=6).border = header_right_border()

    ws.cell(row=4, column=7, value='CHK BY:').font = FONT_TINY
    ws.cell(row=4, column=7).border = BORDER_ALL
    ws.cell(row=4, column=8).font = FONT_VALUE
    ws.cell(row=4, column=8).border = BORDER_ALL

    ws.cell(row=4, column=9, value='REV:').font = FONT_TINY
    ws.cell(row=4, column=9).border = BORDER_ALL
    ws.cell(row=4, column=10, value=revision).font = FONT_VALUE
    ws.cell(row=4, column=10).border = BORDER_ALL

    # --- Row 5: Tag No. [value] | APP BY: | DATE: ---
    ws.merge_cells(start_row=5, start_column=1, end_row=5, end_column=2)
    ws.cell(row=5, column=1, value='Tag No.').font = FONT_LABEL
    ws.cell(row=5, column=1).border = header_left_border()
    ws.cell(row=5, column=2).border = header_right_border()

    ws.merge_cells(start_row=5, start_column=3, end_row=5, end_column=6)
    ws.cell(row=5, column=3, value=equipment_tag).font = FONT_VALUE
    ws.cell(row=5, column=3).border = header_left_border()
    for col in range(4, 6):
        ws.cell(row=5, column=col).border = header_middle_border()
    ws.cell(row=5, column=6).border = header_right_border()

    ws.cell(row=5, column=7, value='APP BY:').font = FONT_TINY
    ws.cell(row=5, column=7).border = BORDER_ALL
    ws.cell(row=5, column=8).font = FONT_VALUE
    ws.cell(row=5, column=8).border = BORDER_ALL

    ws.cell(row=5, column=9, value='DATE:').font = FONT_TINY
    ws.cell(row=5, column=9).border = BORDER_ALL
    ws.cell(row=5, column=10, value=doc_date).font = FONT_VALUE
    ws.cell(row=5, column=10).border = BORDER_ALL

    # --- Row 6: Application: [value] | (empty right side) ---
    ws.merge_cells(start_row=6, start_column=1, end_row=6, end_column=2)
    ws.cell(row=6, column=1, value='Application:').font = FONT_LABEL
    ws.cell(row=6, column=1).border = header_left_border()
    ws.cell(row=6, column=2).border = header_right_border()

    ws.merge_cells(start_row=6, start_column=3, end_row=6, end_column=6)
    ws.cell(row=6, column=3, value=service_type).font = FONT_LABEL
    ws.cell(row=6, column=3).border = header_left_border()
    for col in range(4, 6):
        ws.cell(row=6, column=col).border = header_middle_border()
    ws.cell(row=6, column=6).border = header_right_border()

    for col in range(7, 11):
        ws.cell(row=6, column=col).border = BORDER_ALL

    # --- Row 7: P&ID No: [value] | (empty right side) ---
    ws.merge_cells(start_row=7, start_column=1, end_row=7, end_column=2)
    ws.cell(row=7, column=1, value='P&ID No:').font = FONT_LABEL
    ws.cell(row=7, column=1).border = header_left_border()
    ws.cell(row=7, column=2).border = header_right_border()

    ws.merge_cells(start_row=7, start_column=3, end_row=7, end_column=6)
    ws.cell(row=7, column=3, value=p_and_id).font = FONT_VALUE
    ws.cell(row=7, column=3).border = header_left_border()
    for col in range(4, 6):
        ws.cell(row=7, column=col).border = header_middle_border()
    ws.cell(row=7, column=6).border = header_right_border()

    for col in range(7, 11):
        ws.cell(row=7, column=col).border = BORDER_ALL

    row = 8

    # === OPERATING / DESIGN DATA ===
    operating_data = sections.get('Operating / Design Data', [])
    if operating_data:
        row = write_data_section(ws, row, 'OPERATING / DESIGN DATA', operating_data, validations)

    # === MATERIALS ===
    materials = sections.get('Materials', [])
    if materials:
        row = write_materials_section(ws, row, materials)

    # === DRIVER / MOTOR DATA ===
    has_motor = metadata.get('has_motor', False)
    motor_data = sections.get('Driver / Motor Data', [])
    if has_motor and motor_data:
        row = write_data_section(ws, row, 'DRIVER / MOTOR DATA', motor_data, validations)

    # === REMARKS / NOTES ===
    remarks = sections.get('Remarks', [])
    row = write_remarks_section(ws, row, remarks)

    # === REVISION HISTORY ===
    rev_history = sections.get('Revision History', [])
    row = write_revision_section(ws, row, rev_history)

    # === FOOTER SECTION ===
    row = write_footer_section(ws, row, equipment_tag)

    # Add data validations
    for validation in validations:
        ws.add_data_validation(validation)

    # Apply print layout
    apply_print_layout(ws)

    return wb


def write_data_section(ws, start_row: int, section_title: str, data: list[dict],
                       validations: list) -> int:
    """Write a data section (Operating Data or Motor Data) with professional borders."""
    row = start_row
    section_start = start_row

    # Section header
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=10)
    ws.cell(row=row, column=1, value=section_title).font = FONT_SECTION
    ws.cell(row=row, column=1).alignment = ALIGN_LEFT
    apply_borders_to_range(ws, row, row, 1, 10)
    row += 1

    # Count valid data rows for border calculation
    valid_items = [item for item in data
                   if item.get('#', '') != '#' and item.get('Field', '') != 'Field']

    # Data rows
    for idx, item in enumerate(valid_items):
        row_num = item.get('#', '')
        field_label = item.get('Field', '')
        value = item.get('Value', '')
        units = item.get('Units', '')
        field_type = item.get('Type', 'text')
        options = item.get('Options', '')

        is_last_row = (idx == len(valid_items) - 1)

        # Determine border style: solid on outer edges, dotted internal
        left_border = THIN_BORDER
        right_border = THIN_BORDER
        bottom_border = THIN_BORDER if is_last_row else DOT_BORDER

        # Row number (column A)
        ws.cell(row=row, column=1, value=row_num).font = FONT_SMALL
        ws.cell(row=row, column=1).alignment = ALIGN_CENTER
        ws.cell(row=row, column=1).border = Border(
            left=left_border, right=DOT_BORDER, bottom=bottom_border
        )

        # Label (columns B-D)
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=4)
        ws.cell(row=row, column=2, value=field_label).font = FONT_LABEL
        ws.cell(row=row, column=2).alignment = ALIGN_LEFT
        for col in range(2, 5):
            r_border = DOT_BORDER if col < 4 else DOT_BORDER
            ws.cell(row=row, column=col).border = Border(
                right=r_border, bottom=bottom_border
            )

        # Value (columns E-H)
        ws.merge_cells(start_row=row, start_column=5, end_row=row, end_column=8)
        ws.cell(row=row, column=5, value=value).font = FONT_VALUE
        ws.cell(row=row, column=5).alignment = ALIGN_LEFT
        for col in range(5, 9):
            r_border = DOT_BORDER if col < 8 else DOT_BORDER
            ws.cell(row=row, column=col).border = Border(
                right=r_border, bottom=bottom_border
            )

        # Add dropdown validation if needed
        if field_type == 'dropdown' and options:
            option_list = parse_options(options)
            if option_list:
                # Create validation formula
                formula = f'"{",".join(option_list)}"'
                if len(formula) <= 255:
                    dv = DataValidation(type='list', formula1=formula, allow_blank=True)
                    dv.add(ws.cell(row=row, column=5))
                    validations.append(dv)

        # Units (column I)
        ws.cell(row=row, column=9, value=units).font = FONT_SMALL
        ws.cell(row=row, column=9).alignment = ALIGN_LEFT
        ws.cell(row=row, column=9).border = Border(
            right=DOT_BORDER, bottom=bottom_border
        )

        # Notes (column J)
        ws.cell(row=row, column=10).font = FONT_SMALL
        ws.cell(row=row, column=10).border = Border(
            right=right_border, bottom=bottom_border
        )

        row += 1

    return row


def write_materials_section(ws, start_row: int, materials: list[dict]) -> int:
    """Write the materials section with professional borders."""
    row = start_row

    # Section header
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=10)
    ws.cell(row=row, column=1, value='MATERIALS').font = FONT_SECTION
    ws.cell(row=row, column=1).alignment = ALIGN_LEFT
    apply_borders_to_range(ws, row, row, 1, 10)
    row += 1

    # Column headers with solid bottom border
    ws.cell(row=row, column=1).font = FONT_SMALL
    ws.cell(row=row, column=1).border = Border(
        left=THIN_BORDER, right=DOT_BORDER, bottom=THIN_BORDER
    )

    ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=5)
    ws.cell(row=row, column=2, value='Component').font = FONT_LABEL
    for col in range(2, 6):
        ws.cell(row=row, column=col).border = Border(
            right=DOT_BORDER, bottom=THIN_BORDER
        )

    ws.merge_cells(start_row=row, start_column=6, end_row=row, end_column=10)
    ws.cell(row=row, column=6, value='Material').font = FONT_LABEL
    for col in range(6, 11):
        r_border = THIN_BORDER if col == 10 else DOT_BORDER
        ws.cell(row=row, column=col).border = Border(
            right=r_border, bottom=THIN_BORDER
        )
    row += 1

    # Count valid material rows
    valid_items = [item for item in materials
                   if item.get('#', '') != '#' and item.get('Component', '') != 'Component']

    # Material rows
    for idx, item in enumerate(valid_items):
        row_num = item.get('#', '')
        component = item.get('Component', '')
        material = item.get('Material', '')

        is_last_row = (idx == len(valid_items) - 1)
        bottom_border = THIN_BORDER if is_last_row else DOT_BORDER

        # Row number
        ws.cell(row=row, column=1, value=row_num).font = FONT_SMALL
        ws.cell(row=row, column=1).alignment = ALIGN_CENTER
        ws.cell(row=row, column=1).border = Border(
            left=THIN_BORDER, right=DOT_BORDER, bottom=bottom_border
        )

        # Component (columns B-E)
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=5)
        ws.cell(row=row, column=2, value=component).font = FONT_LABEL
        for col in range(2, 6):
            ws.cell(row=row, column=col).border = Border(
                right=DOT_BORDER, bottom=bottom_border
            )

        # Material (columns F-J)
        ws.merge_cells(start_row=row, start_column=6, end_row=row, end_column=10)
        ws.cell(row=row, column=6, value=material).font = FONT_VALUE
        for col in range(6, 11):
            r_border = THIN_BORDER if col == 10 else DOT_BORDER
            ws.cell(row=row, column=col).border = Border(
                right=r_border, bottom=bottom_border
            )

        row += 1

    return row


def write_remarks_section(ws, start_row: int, remarks: list[dict]) -> int:
    """Write the remarks/notes section with professional numbered format."""
    row = start_row

    # Section header
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=10)
    ws.cell(row=row, column=1, value='NOTES').font = FONT_SECTION
    ws.cell(row=row, column=1).alignment = ALIGN_LEFT
    apply_borders_to_range(ws, row, row, 1, 10)
    row += 1

    # Notes rows (at least 3)
    num_remarks = max(len(remarks), 3)
    for i in range(num_remarks):
        remark_text = ''
        if i < len(remarks):
            item = remarks[i]
            remark_text = item.get('Remark', '')

        is_last_row = (i == num_remarks - 1)
        bottom_border = THIN_BORDER if is_last_row else DOT_BORDER

        # Row number with closing parenthesis: "1)", "2)", etc.
        ws.cell(row=row, column=1, value=f'{i + 1})').font = FONT_SMALL
        ws.cell(row=row, column=1).alignment = ALIGN_CENTER
        ws.cell(row=row, column=1).border = Border(
            left=THIN_BORDER, right=DOT_BORDER, bottom=bottom_border
        )

        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=10)
        ws.cell(row=row, column=2, value=remark_text).font = FONT_VALUE
        for col in range(2, 11):
            r_border = THIN_BORDER if col == 10 else None
            ws.cell(row=row, column=col).border = Border(
                right=r_border, bottom=bottom_border
            )

        row += 1

    return row


def write_revision_section(ws, start_row: int, revisions: list[dict]) -> int:
    """Write the revision history section with compact single-row format."""
    row = start_row

    # Compact revision rows (3 minimum): "Rev.1: Date:" | [date] | | "DESC:" | [description]
    num_revisions = 3
    for i in range(num_revisions):
        date = ''
        description = ''

        if i < len(revisions):
            item = revisions[i]
            date = item.get('Date', '')
            description = item.get('Description', '')

        # "Rev.N: Date:" label (columns A-B)
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
        ws.cell(row=row, column=1, value=f'Rev.{i + 1}: Date:').font = FONT_SMALL
        ws.cell(row=row, column=1).border = BORDER_ALL
        ws.cell(row=row, column=2).border = BORDER_ALL

        # Date value (column C)
        ws.cell(row=row, column=3, value=date).font = FONT_VALUE
        ws.cell(row=row, column=3).border = BORDER_ALL

        # "DESC:" label (column D)
        ws.cell(row=row, column=4, value='DESC:').font = FONT_SMALL
        ws.cell(row=row, column=4).border = BORDER_ALL

        # Description (columns E-J)
        ws.merge_cells(start_row=row, start_column=5, end_row=row, end_column=10)
        ws.cell(row=row, column=5, value=description).font = FONT_VALUE
        for col in range(5, 11):
            ws.cell(row=row, column=col).border = BORDER_ALL

        row += 1

    return row


def write_footer_section(ws, start_row: int, equipment_tag: str) -> int:
    """Write the footer section with Date Released, Spec Status, and tag repeat."""
    row = start_row

    # Blank separator row
    for col in range(1, 11):
        ws.cell(row=row, column=col).border = Border(left=THIN_BORDER if col == 1 else None,
                                                      right=THIN_BORDER if col == 10 else None)
    row += 1

    # Date Released row
    ws.cell(row=row, column=1).border = Border(left=THIN_BORDER)
    ws.cell(row=row, column=2, value='Date Released:').font = FONT_LABEL
    ws.merge_cells(start_row=row, start_column=3, end_row=row, end_column=4)
    ws.cell(row=row, column=3).font = FONT_VALUE
    for col in range(5, 10):
        ws.cell(row=row, column=col).border = Border()
    ws.cell(row=row, column=10).border = Border(right=THIN_BORDER)
    row += 1

    # Spec Status row
    ws.cell(row=row, column=1).border = Border(left=THIN_BORDER)
    ws.cell(row=row, column=2, value='Spec. Status:').font = FONT_LABEL
    ws.merge_cells(start_row=row, start_column=3, end_row=row, end_column=4)
    ws.cell(row=row, column=3).font = FONT_VALUE
    for col in range(5, 10):
        ws.cell(row=row, column=col).border = Border()
    ws.cell(row=row, column=10).border = Border(right=THIN_BORDER)
    row += 1

    # Footer row with tag number at bottom right
    for col in range(1, 7):
        b_left = THIN_BORDER if col == 1 else None
        ws.cell(row=row, column=col).border = Border(left=b_left, bottom=THIN_BORDER)

    # Tag number at bottom right (bold)
    ws.merge_cells(start_row=row, start_column=7, end_row=row, end_column=10)
    ws.cell(row=row, column=7, value=equipment_tag).font = FONT_VALUE_BOLD
    ws.cell(row=row, column=7).alignment = ALIGN_RIGHT
    for col in range(7, 11):
        r_border = THIN_BORDER if col == 10 else None
        ws.cell(row=row, column=col).border = Border(right=r_border, bottom=THIN_BORDER)

    row += 1
    return row


# =============================================================================
# OUTPUT FILENAME GENERATION
# =============================================================================

def generate_output_filename(template: dict) -> str:
    """Generate the output Excel filename based on template metadata."""
    metadata = template['metadata']
    template_id = metadata.get('template_id', 'UNKNOWN')
    title = metadata.get('title', 'DATASHEET')

    # Format: 101-GR-01 GRIT REMOVAL PACKAGE.xlsx
    return f"{template_id}-01 {title}.xlsx"


# =============================================================================
# MAIN
# =============================================================================

def convert_template(input_path: Path, output_dir: Path) -> bool:
    """Convert a single markdown template to Excel."""
    print(f"Converting: {input_path.name}")

    try:
        template = parse_template(input_path)
        workbook = build_workbook(template)

        output_filename = generate_output_filename(template)
        output_path = output_dir / output_filename

        workbook.save(output_path)
        print(f"  -> {output_path.name}")
        return True

    except Exception as e:
        print(f"  ERROR: {e}")
        import traceback
        traceback.print_exc()
        return False


def main():
    parser = argparse.ArgumentParser(
        description='Convert markdown templates to Excel datasheets'
    )
    parser.add_argument('files', nargs='*', help='Markdown files to convert')
    parser.add_argument('--all', action='store_true',
                        help='Convert all templates in templates/ directory')
    parser.add_argument('--output-dir', '-o', default='assets',
                        help='Output directory for Excel files (default: assets)')

    args = parser.parse_args()

    # Determine script location for relative paths
    script_dir = Path(__file__).parent
    project_dir = script_dir.parent

    output_dir = Path(args.output_dir)
    if not output_dir.is_absolute():
        output_dir = project_dir / output_dir
    output_dir.mkdir(parents=True, exist_ok=True)

    # Collect files to convert
    if args.all:
        templates_dir = project_dir / 'templates'
        files = list(templates_dir.glob('*.md'))
    elif args.files:
        files = [Path(f) for f in args.files]
    else:
        print("Usage: md_to_excel.py <files...> or --all")
        sys.exit(1)

    if not files:
        print("No template files found")
        sys.exit(1)

    print(f"Converting {len(files)} template(s)...\n")

    success_count = 0
    for filepath in sorted(files):
        if not filepath.exists():
            print(f"Warning: File not found: {filepath}")
            continue

        if convert_template(filepath, output_dir):
            success_count += 1

    print(f"\nCompleted: {success_count}/{len(files)} templates converted successfully")

    if success_count < len(files):
        sys.exit(1)


if __name__ == '__main__':
    main()
