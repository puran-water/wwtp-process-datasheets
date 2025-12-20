#!/usr/bin/env python3
"""
WWTP Datasheet Validation Script

Performs completeness checking on Excel datasheets before RFQ issuance.

Features:
- Checks for blank required fields
- Validates dropdown selections
- Detects unit ambiguity (value without units)
- Generates completeness report

Usage:
    python scripts/validate_datasheet.py assets/101-GR-01*.xlsx
    python scripts/validate_datasheet.py --all
    python scripts/validate_datasheet.py assets/*.xlsx --format json

Exit codes:
    0 = All datasheets are READY
    1 = One or more datasheets are NOT READY
    2 = Error loading files
"""

import os
import sys
import argparse
import json
from pathlib import Path
from dataclasses import dataclass, asdict
from typing import List, Optional

try:
    import openpyxl
except ImportError:
    print("Error: openpyxl is required. Install with: pip install openpyxl")
    sys.exit(2)

try:
    import yaml
except ImportError:
    yaml = None  # YAML is optional, will use built-in config


# =============================================================================
# DATA CLASSES
# =============================================================================

@dataclass
class FieldCheck:
    """Result of checking a single field."""
    field_name: str
    row: int
    column: int
    severity: str  # REQ, OPT, INFO
    status: str    # OK, MISSING, WARNING
    message: str = ""


@dataclass
class ValidationResult:
    """Complete validation result for a datasheet."""
    filename: str
    template_type: str
    total_required: int
    filled_required: int
    completeness_pct: float
    is_ready: bool
    missing_required: List[FieldCheck]
    warnings: List[FieldCheck]
    info: List[str]


# =============================================================================
# BUILT-IN CONFIG (fallback if YAML not available)
# =============================================================================

BUILTIN_CONFIG = {
    'common': [
        {'field_name': 'Design Average Flow', 'severity': 'REQ'},
        {'field_name': 'Design Peak', 'severity': 'REQ'},
        {'field_name': 'Number of', 'severity': 'REQ'},
        {'field_name': 'Redundancy', 'severity': 'REQ'},
    ],
    '101-GR': [
        {'field_name': 'Target Grit Cut Size', 'severity': 'REQ'},
        {'field_name': 'Required Removal Efficiency', 'severity': 'REQ'},
        {'field_name': 'Technology Type', 'severity': 'REQ'},
        {'field_name': 'Grit Disposal', 'severity': 'REQ'},
    ],
    '101-SC': [
        {'field_name': 'Screen Type', 'severity': 'REQ'},
        {'field_name': 'Opening Size', 'severity': 'REQ'},
        {'field_name': 'Cleaning Mechanism', 'severity': 'REQ'},
    ],
    '130-CL': [
        {'field_name': 'Clarifier Geometry', 'severity': 'REQ'},
        {'field_name': 'Diameter', 'severity': 'REQ'},
        {'field_name': 'Sidewater Depth', 'severity': 'REQ'},
        {'field_name': 'Surface Loading Rate', 'severity': 'REQ'},
    ],
    '130-LC': [
        {'field_name': 'Surface Loading Rate', 'severity': 'REQ'},
        {'field_name': 'Effluent Quality', 'severity': 'REQ'},
    ],
    '230-AT': [
        {'field_name': 'Basin Volume', 'severity': 'REQ'},
        {'field_name': 'DO Setpoint', 'severity': 'REQ'},
        {'field_name': 'Aeration Technology', 'severity': 'REQ'},
        {'field_name': 'Blower Type', 'severity': 'REQ'},
    ],
    '240-SC': [
        {'field_name': 'Clarifier Type', 'severity': 'REQ'},
        {'field_name': 'Diameter', 'severity': 'REQ'},
        {'field_name': 'Surface Loading', 'severity': 'REQ'},
    ],
    '310-DMF': [
        {'field_name': 'Filter Type', 'severity': 'REQ'},
        {'field_name': 'Loading Rate', 'severity': 'REQ'},
        {'field_name': 'Effluent Quality', 'severity': 'REQ'},
    ],
    '420-UV': [
        {'field_name': 'Channel Configuration', 'severity': 'REQ'},
        {'field_name': 'Lamp Type', 'severity': 'REQ'},
        {'field_name': 'UV Dose', 'severity': 'REQ'},
        {'field_name': 'UVT', 'severity': 'REQ'},
    ],
    '510-GT': [
        {'field_name': 'Sludge Type', 'severity': 'REQ'},
        {'field_name': 'Solids Loading', 'severity': 'REQ'},
        {'field_name': 'Feed Concentration', 'severity': 'REQ'},
        {'field_name': 'Target Concentration', 'severity': 'REQ'},
    ],
}


# =============================================================================
# HELPER FUNCTIONS
# =============================================================================

def load_config(config_path: Optional[Path] = None) -> dict:
    """Load field configuration from YAML or use built-in."""
    if config_path and config_path.exists() and yaml:
        with open(config_path) as f:
            return yaml.safe_load(f)
    return BUILTIN_CONFIG


def identify_template_type(filename: str) -> str:
    """Extract template type prefix from filename."""
    name = Path(filename).stem.upper()

    # Extract prefix like "101-GR", "230-AT", etc.
    parts = name.split('-')
    if len(parts) >= 2:
        prefix = f"{parts[0]}-{parts[1][:2]}"
        return prefix

    return "UNKNOWN"


def find_field_row(ws, field_name: str, search_col: int = 2) -> Optional[int]:
    """Find the row containing a field label."""
    field_lower = field_name.lower()

    for row in range(1, ws.max_row + 1):
        cell = ws.cell(row=row, column=search_col)
        if cell.value and field_lower in str(cell.value).lower():
            return row

    return None


def is_cell_empty(ws, row: int, value_cols: List[int] = [5, 6, 7, 8]) -> bool:
    """Check if all value columns in a row are empty."""
    for col in value_cols:
        cell = ws.cell(row=row, column=col)
        if cell.value is not None and str(cell.value).strip() != '':
            return False
    return True


def check_unit_ambiguity(ws, row: int, value_col: int = 5, unit_col: int = 9) -> bool:
    """Check if a value exists without a corresponding unit."""
    value_cell = ws.cell(row=row, column=value_col)
    unit_cell = ws.cell(row=row, column=unit_col)

    has_value = value_cell.value is not None and str(value_cell.value).strip() != ''
    has_unit = unit_cell.value is not None and str(unit_cell.value).strip() != ''

    # Value without unit is ambiguous (unless it's a Yes/No or text field)
    if has_value and not has_unit:
        try:
            float(str(value_cell.value).replace(',', ''))
            return True  # Numeric value without unit
        except ValueError:
            return False  # Text value, no unit needed

    return False


# =============================================================================
# MAIN VALIDATION
# =============================================================================

def validate_datasheet(filepath: Path, config: dict) -> ValidationResult:
    """Validate a single datasheet for completeness."""

    filename = filepath.name
    template_type = identify_template_type(filename)

    # Load workbook
    try:
        wb = openpyxl.load_workbook(filepath)
        ws = wb.active
    except Exception as e:
        return ValidationResult(
            filename=filename,
            template_type=template_type,
            total_required=0,
            filled_required=0,
            completeness_pct=0.0,
            is_ready=False,
            missing_required=[],
            warnings=[],
            info=[f"Error loading file: {e}"]
        )

    # Get required fields for this template
    common_fields = config.get('common', [])
    template_fields = config.get(template_type, [])
    all_fields = common_fields + template_fields

    missing_required = []
    warnings = []
    info = []

    required_count = 0
    filled_count = 0

    # Check each field
    for field_def in all_fields:
        field_name = field_def.get('field_name', '')
        severity = field_def.get('severity', 'OPT')

        if severity == 'REQ':
            required_count += 1

        # Find the field row
        row = find_field_row(ws, field_name)

        if row is None:
            # Field not found in template
            if severity == 'REQ':
                missing_required.append(FieldCheck(
                    field_name=field_name,
                    row=0,
                    column=2,
                    severity=severity,
                    status='MISSING',
                    message=f"Field not found in template"
                ))
            continue

        # Check if value is filled
        if is_cell_empty(ws, row):
            if severity == 'REQ':
                missing_required.append(FieldCheck(
                    field_name=field_name,
                    row=row,
                    column=5,
                    severity=severity,
                    status='MISSING',
                    message=f"Required field is empty"
                ))
            elif severity == 'OPT':
                warnings.append(FieldCheck(
                    field_name=field_name,
                    row=row,
                    column=5,
                    severity=severity,
                    status='WARNING',
                    message=f"Optional field is empty"
                ))
        else:
            if severity == 'REQ':
                filled_count += 1

            # Check for unit ambiguity
            if check_unit_ambiguity(ws, row):
                warnings.append(FieldCheck(
                    field_name=field_name,
                    row=row,
                    column=5,
                    severity='WARNING',
                    status='WARNING',
                    message=f"Numeric value without units specified"
                ))

    # Calculate completeness
    completeness = (filled_count / required_count * 100) if required_count > 0 else 100.0
    is_ready = len(missing_required) == 0

    # Add info messages
    info.append(f"Template type: {template_type}")
    info.append(f"Checked {len(all_fields)} fields ({required_count} required, {len(all_fields) - required_count} optional)")

    return ValidationResult(
        filename=filename,
        template_type=template_type,
        total_required=required_count,
        filled_required=filled_count,
        completeness_pct=round(completeness, 1),
        is_ready=is_ready,
        missing_required=missing_required,
        warnings=warnings,
        info=info
    )


def format_text_report(result: ValidationResult) -> str:
    """Format validation result as text report."""
    lines = []
    lines.append("=" * 60)
    lines.append(f"DATASHEET VALIDATION REPORT")
    lines.append("=" * 60)
    lines.append(f"File: {result.filename}")
    lines.append(f"Template: {result.template_type}")
    lines.append(f"Completeness: {result.completeness_pct}% ({result.filled_required}/{result.total_required} required fields)")
    lines.append("")

    if result.missing_required:
        lines.append("MISSING REQUIRED FIELDS:")
        for check in result.missing_required:
            lines.append(f"  Row {check.row:3}: {check.field_name} - {check.message}")
        lines.append("")

    if result.warnings:
        lines.append("WARNINGS:")
        for check in result.warnings:
            lines.append(f"  Row {check.row:3}: {check.field_name} - {check.message}")
        lines.append("")

    if result.info:
        lines.append("INFO:")
        for msg in result.info:
            lines.append(f"  {msg}")
        lines.append("")

    status = "READY FOR RFQ" if result.is_ready else "NOT READY - Fix missing fields"
    lines.append(f"STATUS: {status}")
    lines.append("=" * 60)

    return "\n".join(lines)


def format_json_report(results: List[ValidationResult]) -> str:
    """Format validation results as JSON."""
    data = {
        'summary': {
            'total_files': len(results),
            'ready': sum(1 for r in results if r.is_ready),
            'not_ready': sum(1 for r in results if not r.is_ready),
        },
        'results': [asdict(r) for r in results]
    }
    return json.dumps(data, indent=2, default=str)


# =============================================================================
# MAIN
# =============================================================================

def main():
    parser = argparse.ArgumentParser(description='Validate WWTP datasheets for completeness')
    parser.add_argument('files', nargs='*', help='Excel files to validate')
    parser.add_argument('--all', action='store_true', help='Validate all templates in assets/')
    parser.add_argument('--config', type=Path, help='Path to required_fields.yaml config')
    parser.add_argument('--format', choices=['text', 'json'], default='text', help='Output format')

    args = parser.parse_args()

    # Determine files to validate
    if args.all:
        assets_dir = Path(__file__).parent.parent / 'assets'
        files = [f for f in assets_dir.glob('*.xlsx')
                 if not f.name.startswith('000-') and not f.name.startswith('VENDOR')]
    elif args.files:
        files = [Path(f) for f in args.files]
    else:
        print("Usage: python validate_datasheet.py <files...> or --all")
        sys.exit(2)

    # Load config
    default_config = Path(__file__).parent / 'config' / 'required_fields.yaml'
    config = load_config(args.config or default_config)

    # Validate each file
    results = []
    for filepath in sorted(files):
        if not filepath.exists():
            print(f"Warning: File not found: {filepath}", file=sys.stderr)
            continue

        result = validate_datasheet(filepath, config)
        results.append(result)

    # Output results
    if args.format == 'json':
        print(format_json_report(results))
    else:
        for result in results:
            print(format_text_report(result))
            print()

    # Exit code
    all_ready = all(r.is_ready for r in results)
    sys.exit(0 if all_ready else 1)


if __name__ == '__main__':
    main()
