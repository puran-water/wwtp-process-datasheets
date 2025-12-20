#!/usr/bin/env python3
"""
Centralized Excel Styling Definitions for WWTP Datasheets

Provides consistent, professional styling matching instrument datasheet format:
- Solid black borders (#000000)
- Font hierarchy (12pt title, 10pt body, 8pt labels)
- No color fills (monochrome)
- Single-page print layout

Uses NamedStyle objects to prevent Excel style explosion.
"""

from openpyxl.styles import (
    Font, Border, Side, Alignment, PatternFill, NamedStyle
)
from openpyxl.styles.colors import Color


# =============================================================================
# COLOR DEFINITIONS
# =============================================================================

# Solid black - use ARGB format (FF = fully opaque)
BLACK = Color(rgb='FF000000')

# No fill (white background)
NO_FILL = PatternFill(fill_type=None)


# =============================================================================
# BORDER DEFINITIONS
# =============================================================================

THIN_BORDER = Side(style='thin', color=BLACK)

BORDER_ALL = Border(
    left=THIN_BORDER,
    right=THIN_BORDER,
    top=THIN_BORDER,
    bottom=THIN_BORDER
)

BORDER_NONE = Border()


# =============================================================================
# FONT DEFINITIONS
# =============================================================================

# Title: 12pt bold
FONT_TITLE = Font(name='Arial', size=12, bold=True)

# Section header: 10pt bold
FONT_SECTION = Font(name='Arial', size=10, bold=True)

# Label: 10pt regular
FONT_LABEL = Font(name='Arial', size=10)

# Value: 10pt regular
FONT_VALUE = Font(name='Arial', size=10)

# Small: 8pt regular (row numbers, units, notes)
FONT_SMALL = Font(name='Arial', size=8)

# Default: 11pt (Excel default, used for unstyled cells)
FONT_DEFAULT = Font(name='Arial', size=11)


# =============================================================================
# ALIGNMENT DEFINITIONS
# =============================================================================

ALIGN_LEFT = Alignment(horizontal='left', vertical='center', wrap_text=True)
ALIGN_CENTER = Alignment(horizontal='center', vertical='center')
ALIGN_RIGHT = Alignment(horizontal='right', vertical='center')
ALIGN_TOP_LEFT = Alignment(horizontal='left', vertical='top', wrap_text=True)


# =============================================================================
# COLUMN LAYOUT (10 columns: A-J)
# =============================================================================

COLUMN_WIDTHS = {
    'A': 4,    # Row number
    'B': 24,   # Label (part 1)
    'C': 10,   # Label (part 2)
    'D': 10,   # Label (part 3)
    'E': 12,   # Value (part 1)
    'F': 10,   # Value (part 2)
    'G': 10,   # Value (part 3)
    'H': 10,   # Value (part 4)
    'I': 12,   # Units
    'J': 14,   # Notes/Vendor
}

# Row height in points (fits ~55-60 rows on single page)
ROW_HEIGHT = 15


# =============================================================================
# NAMED STYLES (Reusable, prevent style explosion)
# =============================================================================

def create_named_styles():
    """Create NamedStyle objects for efficient styling."""
    styles = {}

    # Title style
    title_style = NamedStyle(name='datasheet_title')
    title_style.font = FONT_TITLE
    title_style.alignment = ALIGN_CENTER
    title_style.border = BORDER_ALL
    title_style.fill = NO_FILL
    styles['title'] = title_style

    # Section header style
    section_style = NamedStyle(name='datasheet_section')
    section_style.font = FONT_SECTION
    section_style.alignment = ALIGN_LEFT
    section_style.border = BORDER_ALL
    section_style.fill = NO_FILL
    styles['section'] = section_style

    # Label style
    label_style = NamedStyle(name='datasheet_label')
    label_style.font = FONT_LABEL
    label_style.alignment = ALIGN_LEFT
    label_style.border = BORDER_ALL
    label_style.fill = NO_FILL
    styles['label'] = label_style

    # Value style
    value_style = NamedStyle(name='datasheet_value')
    value_style.font = FONT_VALUE
    value_style.alignment = ALIGN_LEFT
    value_style.border = BORDER_ALL
    value_style.fill = NO_FILL
    styles['value'] = value_style

    # Small text style (row numbers, units)
    small_style = NamedStyle(name='datasheet_small')
    small_style.font = FONT_SMALL
    small_style.alignment = ALIGN_CENTER
    small_style.border = BORDER_ALL
    small_style.fill = NO_FILL
    styles['small'] = small_style

    # Default cell style
    default_style = NamedStyle(name='datasheet_default')
    default_style.font = FONT_DEFAULT
    default_style.alignment = ALIGN_LEFT
    default_style.border = BORDER_ALL
    default_style.fill = NO_FILL
    styles['default'] = default_style

    return styles


def register_styles(workbook):
    """Register named styles with a workbook (call once per workbook)."""
    styles = create_named_styles()
    for style in styles.values():
        try:
            workbook.add_named_style(style)
        except ValueError:
            # Style already exists
            pass
    return styles


# =============================================================================
# PRINT LAYOUT SETTINGS
# =============================================================================

def apply_print_layout(worksheet):
    """Configure worksheet for single-page print layout."""
    ws = worksheet

    # Page setup
    ws.page_setup.orientation = 'portrait'
    ws.page_setup.paperSize = ws.PAPERSIZE_LETTER

    # Narrow margins (0.25 inches)
    ws.page_margins.left = 0.25
    ws.page_margins.right = 0.25
    ws.page_margins.top = 0.25
    ws.page_margins.bottom = 0.25
    ws.page_margins.header = 0.1
    ws.page_margins.footer = 0.1

    # Fit to one page
    try:
        ws.page_setup.fitToPage = True
        ws.page_setup.fitToWidth = 1
        ws.page_setup.fitToHeight = 1
    except Exception:
        pass  # Some openpyxl versions don't support all settings


def apply_column_widths(worksheet):
    """Set column widths for 10-column layout."""
    for col_letter, width in COLUMN_WIDTHS.items():
        worksheet.column_dimensions[col_letter].width = width


def apply_row_heights(worksheet, start_row=1, end_row=60):
    """Set consistent row heights for single-page fit."""
    for row_num in range(start_row, end_row + 1):
        worksheet.row_dimensions[row_num].height = ROW_HEIGHT


# =============================================================================
# CELL STYLING HELPERS
# =============================================================================

def style_cell(cell, style_type='default'):
    """Apply styling to a single cell."""
    if style_type == 'title':
        cell.font = FONT_TITLE
        cell.alignment = ALIGN_CENTER
    elif style_type == 'section':
        cell.font = FONT_SECTION
        cell.alignment = ALIGN_LEFT
    elif style_type == 'label':
        cell.font = FONT_LABEL
        cell.alignment = ALIGN_LEFT
    elif style_type == 'value':
        cell.font = FONT_VALUE
        cell.alignment = ALIGN_LEFT
    elif style_type == 'small':
        cell.font = FONT_SMALL
        cell.alignment = ALIGN_CENTER
    else:
        cell.font = FONT_DEFAULT
        cell.alignment = ALIGN_LEFT

    cell.border = BORDER_ALL
    cell.fill = NO_FILL


def apply_borders_to_range(worksheet, start_row, end_row, start_col, end_col):
    """Apply solid black borders to all cells in a range."""
    for row in range(start_row, end_row + 1):
        for col in range(start_col, end_col + 1):
            cell = worksheet.cell(row=row, column=col)
            cell.border = BORDER_ALL
            cell.fill = NO_FILL


# =============================================================================
# TEMPLATE LAYOUT CONSTANTS
# =============================================================================

# Excel layout structure
LAYOUT = {
    'title_rows': (1, 2),           # Merged title block
    'header_rows': (3, 8),          # Service info + document control
    'data_start_row': 9,            # First data section starts here
    'total_columns': 10,            # A through J
    'label_cols': (2, 4),           # B-D for labels
    'value_cols': (5, 8),           # E-H for values
    'units_col': 9,                 # I for units
    'notes_col': 10,                # J for notes
}
