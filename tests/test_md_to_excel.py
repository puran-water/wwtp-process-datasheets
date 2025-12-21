#!/usr/bin/env python3
"""
Tests for md_to_excel.py

Run with: pytest tests/test_md_to_excel.py -v
"""

import sys
import tempfile
from pathlib import Path

# Add scripts directory to path for imports
sys.path.insert(0, str(Path(__file__).parent.parent / 'scripts'))

import pytest

from md_to_excel import (
    parse_frontmatter,
    parse_markdown_table,
    extract_sections,
    parse_template,
    parse_options,
)


FIXTURES_DIR = Path(__file__).parent / 'fixtures'


class TestParseFrontmatter:
    """Tests for YAML frontmatter parsing."""

    def test_valid_frontmatter(self):
        content = """---
schema_version: 1
template_id: TEST-01
title: TEST EQUIPMENT
service: TEST SERVICE
has_motor: true
---
# Body content
"""
        fm, remaining = parse_frontmatter(content)

        assert fm['schema_version'] == 1
        assert fm['template_id'] == 'TEST-01'
        assert fm['title'] == 'TEST EQUIPMENT'
        assert fm['has_motor'] is True
        assert '# Body content' in remaining

    def test_missing_frontmatter(self):
        content = "# No frontmatter here"
        fm, remaining = parse_frontmatter(content)

        assert fm == {}
        assert remaining == content

    def test_unclosed_frontmatter(self):
        content = """---
schema_version: 1
# Missing closing ---
"""
        fm, remaining = parse_frontmatter(content)
        assert fm == {}


class TestParseMarkdownTable:
    """Tests for markdown table parsing."""

    def test_basic_table(self):
        table = """| Field | Value |
|-------|-------|
| Name | Test |
| Size | 100 |
"""
        rows = parse_markdown_table(table)

        assert len(rows) == 2
        assert rows[0]['Field'] == 'Name'
        assert rows[0]['Value'] == 'Test'
        assert rows[1]['Field'] == 'Size'
        assert rows[1]['Value'] == '100'

    def test_data_table(self):
        table = """| # | field_id | Field | Value | Units | Type | Options |
|---|----------|-------|-------|-------|------|---------|
| 1 | flow | Flow Rate | 100 | m3/h | number | |
| 2 | mode | Mode | A | | dropdown | ["A","B"] |
"""
        rows = parse_markdown_table(table)

        assert len(rows) == 2
        assert rows[0]['field_id'] == 'flow'
        assert rows[0]['Type'] == 'number'
        assert rows[1]['field_id'] == 'mode'
        assert rows[1]['Options'] == '["A","B"]'

    def test_empty_table(self):
        rows = parse_markdown_table("")
        assert rows == []

    def test_table_with_empty_cells(self):
        table = """| Field | Value |
|-------|-------|
| Name | |
| | Data |
"""
        rows = parse_markdown_table(table)
        # Should handle empty cells
        assert len(rows) >= 1


class TestExtractSections:
    """Tests for section extraction."""

    def test_extract_sections(self):
        content = """# Main Title

## Document Control

Control content here

## Operating / Design Data

Data content here

## Materials

Material content
"""
        sections = extract_sections(content)

        assert 'Document Control' in sections
        assert 'Operating / Design Data' in sections
        assert 'Materials' in sections
        assert 'Control content here' in sections['Document Control']

    def test_empty_content(self):
        sections = extract_sections("")
        assert sections == {}

    def test_no_sections(self):
        content = "Just some text without sections"
        sections = extract_sections(content)
        assert sections == {}


class TestParseTemplate:
    """Tests for full template parsing."""

    def test_parse_valid_template(self):
        filepath = FIXTURES_DIR / 'valid_template.md'
        template = parse_template(filepath)

        # Check metadata
        assert template['metadata']['template_id'] == 'TEST-01'
        assert template['metadata']['title'] == 'TEST EQUIPMENT'
        assert template['metadata']['has_motor'] is True

        # Check sections exist
        assert 'Document Control' in template['sections']
        assert 'Operating / Design Data' in template['sections']
        assert 'Materials' in template['sections']
        assert 'Driver / Motor Data' in template['sections']

        # Check data parsed
        op_data = template['sections']['Operating / Design Data']
        assert len(op_data) >= 4
        field_ids = [row.get('field_id') for row in op_data]
        assert 'qavg' in field_ids
        assert 'redundancy' in field_ids


class TestParseOptions:
    """Tests for dropdown options parsing."""

    def test_json_array(self):
        options = parse_options('["A","B","C"]')
        assert options == ['A', 'B', 'C']

    def test_empty_string(self):
        options = parse_options('')
        assert options == []

    def test_none(self):
        options = parse_options(None)
        assert options == []


class TestExcelGeneration:
    """Integration tests for Excel generation."""

    def test_generate_excel_from_template(self):
        """Test that we can generate an Excel file from a valid template."""
        # Import the generation function
        from md_to_excel import generate_datasheet

        filepath = FIXTURES_DIR / 'valid_template.md'
        template = parse_template(filepath)

        with tempfile.TemporaryDirectory() as tmpdir:
            output_path = Path(tmpdir) / 'test_output.xlsx'

            # Generate the Excel file
            generate_datasheet(template, output_path)

            # Verify file was created
            assert output_path.exists()
            assert output_path.stat().st_size > 0

            # Verify it's a valid Excel file
            from openpyxl import load_workbook
            wb = load_workbook(output_path)
            ws = wb.active

            # Check that worksheet has content
            assert ws.max_row > 1
            assert ws.max_column > 1


class TestOutputNaming:
    """Tests for output file naming conventions."""

    def test_output_filename_generation(self):
        """Test that output filenames follow the expected pattern."""
        from md_to_excel import get_output_filename

        # Test with template metadata
        metadata = {
            'template_id': '101-GR',
            'title': 'GRIT REMOVAL PACKAGE'
        }

        filename = get_output_filename(metadata)
        assert '101-GR' in filename
        assert 'GRIT REMOVAL' in filename.upper()
        assert filename.endswith('.xlsx')


if __name__ == '__main__':
    pytest.main([__file__, '-v'])
